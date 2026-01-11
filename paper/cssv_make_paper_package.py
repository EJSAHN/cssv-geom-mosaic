#!/usr/bin/env python3
"""
Make FINAL paper-ready figures (PDF vector + PNG 300 dpi), Table 1 (Excel),
and package key outputs into Supplementary Data S1 (single Excel file with multiple sheets).

Outputs:
- <out_dir>/figures/Figure1...Figure6 (PDF + PNG 300 dpi)
- <out_dir>/tables/Table1_chimera_topN.xlsx (+ .csv)
- <out_dir>/supplementary/Supplementary_Data_S1.xlsx  (multi-sheet Excel)

Notes:
- No raw genomes are packaged; only derived outputs.
- Panel letters can be uppercase (A,B,...) or lowercase (a,b,...).
- Figure styling uses mathtext for italic variables (k, r, p).
"""

from __future__ import annotations

import argparse
import math
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

import matplotlib as mpl
import matplotlib.pyplot as plt

# Illustrator-friendly PDFs (text stays as text)
mpl.rcParams["pdf.fonttype"] = 42
mpl.rcParams["ps.fonttype"] = 42


# --------------------------
# Helpers
# --------------------------
def _ensure_dir(p: Path) -> Path:
    p.mkdir(parents=True, exist_ok=True)
    return p


def _read_square_matrix_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, index_col=0)
    df.index = df.index.map(lambda x: str(x).strip())
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _tip_to_accession(name: str) -> str:
    s = str(name).strip()
    s = s.split("|")[0]
    s = s.split("__dup")[0]
    return s


def _panel_label(
    ax: plt.Axes,
    label: str,
    case: str = "upper",
    xy: Tuple[float, float] = (0.01, 0.99),
    fontsize: float = 14.0,
) -> None:
    label = label.lower() if case.lower().startswith("lower") else label.upper()
    ax.text(
        xy[0],
        xy[1],
        label,
        transform=ax.transAxes,
        ha="left",
        va="top",
        fontsize=fontsize,
        fontweight="bold",
    )


def _save_fig(fig: plt.Figure, out_pdf: Path, out_png: Path, dpi: int = 300) -> None:
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_pdf, bbox_inches="tight")
    fig.savefig(out_png, dpi=dpi, bbox_inches="tight")
    plt.close(fig)


def _upper_tri_vec(m: np.ndarray) -> np.ndarray:
    iu = np.triu_indices(m.shape[0], k=1)
    return m[iu]


# --------------------------
# Figure builders
# --------------------------
def make_figure1(
    genome_summary_csv: Path,
    genome_embedding_csv: Path,
    kmer_dist_csv: Path,
    out_dir: Path,
    panel_case: str = "upper",
    dpi: int = 300,
) -> Tuple[Path, Path]:
    """
    Figure 1:
      (A) length vs GC (label ALL points for consistency)
      (B) MDS embedding (label ALL points)
      (C) k-mer cosine distance heatmap (ordered by MDS1) with improved aspect
    """
    gs = pd.read_csv(genome_summary_csv)
    emb = pd.read_csv(genome_embedding_csv)
    D = _read_square_matrix_csv(kmer_dist_csv)

    # Normalize names
    gs["name"] = gs["name"].astype(str).map(_tip_to_accession)
    emb["name"] = emb["name"].astype(str).map(_tip_to_accession)
    D.index = [_tip_to_accession(x) for x in D.index]
    D.columns = [_tip_to_accession(x) for x in D.columns]

    # Order by MDS1 for heatmap (deterministic)
    emb2 = emb.dropna(subset=["mds1", "mds2"]).copy()
    emb2 = emb2.sort_values(["mds1", "mds2"]).reset_index(drop=True)
    order = [n for n in emb2["name"].tolist() if n in D.index]
    if len(order) >= 3:
        D = D.loc[order, order]

    fig = plt.figure(figsize=(14, 10))
    grid = fig.add_gridspec(nrows=2, ncols=2, height_ratios=[1.0, 1.2], width_ratios=[1.0, 1.15])
    axA = fig.add_subplot(grid[0, 0])
    axB = fig.add_subplot(grid[0, 1])
    axC = fig.add_subplot(grid[1, :])

    # (A) length vs GC
    axA.scatter(gs["length"].to_numpy(dtype=float), gs["gc"].to_numpy(dtype=float), s=30)
    axA.set_xlabel("Genome length (bp)")
    axA.set_ylabel(r"$GC$ fraction")
    axA.set_title("QC: length vs GC")
    # label ALL points (consistent)
    for _, r in gs.iterrows():
        axA.text(float(r["length"]), float(r["gc"]), str(r["name"]), fontsize=5, alpha=0.85, ha="left", va="bottom")
    _panel_label(axA, "A", case=panel_case)

    # (B) MDS
    axB.scatter(emb2["mds1"], emb2["mds2"], s=30)
    axB.set_xlabel("MDS1")
    axB.set_ylabel("MDS2")
    axB.set_title(r"Metric MDS from $k$-mer cosine distances ($k=4$)")
    for _, r in emb2.iterrows():
        axB.text(float(r["mds1"]), float(r["mds2"]), str(r["name"]), fontsize=5, alpha=0.85, ha="left", va="bottom")
    _panel_label(axB, "B", case=panel_case)

    # (C) heatmap (improve aspect + labels)
    mat = D.to_numpy(dtype=float)
    axC.imshow(mat, aspect="equal", interpolation="nearest")
    axC.set_title(r"Genome–genome $k$-mer cosine distance heatmap ($k=4$)")
    axC.set_xlabel("Genomes (ordered by MDS1)")
    axC.set_ylabel("Genomes (ordered by MDS1)")
    axC.set_xticks(range(len(D.columns)))
    axC.set_yticks(range(len(D.index)))
    axC.set_xticklabels(D.columns, rotation=90, fontsize=4)
    axC.set_yticklabels(D.index, fontsize=4)
    cbar = fig.colorbar(axC.images[0], ax=axC, fraction=0.02, pad=0.01)
    cbar.set_label(r"Cosine distance ($k$-mer freq)")
    _panel_label(axC, "C", case=panel_case)

    fig.tight_layout()
    out_pdf = out_dir / "Figure1_genome_geometry.pdf"
    out_png = out_dir / "Figure1_genome_geometry.png"
    _save_fig(fig, out_pdf, out_png, dpi=dpi)
    return out_pdf, out_png


def _normalize_switchpoints(sw: pd.DataFrame) -> pd.DataFrame:
    sw = sw.copy()
    cols = set(sw.columns)

    if {"prev_label", "new_label"}.issubset(cols) and not {"from_label", "to_label"}.issubset(cols):
        sw = sw.rename(columns={"prev_label": "from_label", "new_label": "to_label"})

    if "run_len" not in sw.columns:
        if {"prev_run_windows", "new_run_windows"}.issubset(cols):
            sw["run_len"] = np.minimum(
                sw["prev_run_windows"].astype(float),
                sw["new_run_windows"].astype(float),
            )
        elif "run_windows" in sw.columns:
            sw["run_len"] = sw["run_windows"].astype(float)
        else:
            sw["run_len"] = np.nan

    required = {"name", "pos0", "from_label", "to_label", "run_len"}
    missing = sorted(list(required - set(sw.columns)))
    if missing:
        raise ValueError(f"switchpoints.csv missing columns: {missing} ; found={list(sw.columns)}")

    sw["name"] = sw["name"].astype(str).map(_tip_to_accession)
    sw["pos0"] = pd.to_numeric(sw["pos0"], errors="coerce").fillna(0).astype(int)
    sw["from_label"] = pd.to_numeric(sw["from_label"], errors="coerce").fillna(0).astype(int)
    sw["to_label"] = pd.to_numeric(sw["to_label"], errors="coerce").fillna(0).astype(int)
    sw["run_len"] = pd.to_numeric(sw["run_len"], errors="coerce")
    return sw


def _switchpoint_transition_matrix(sw: pd.DataFrame) -> pd.DataFrame:
    counts = sw.groupby(["from_label", "to_label"]).size().rename("count").reset_index()
    mat = counts.pivot_table(index="from_label", columns="to_label", values="count", fill_value=0)
    mat = mat.sort_index(axis=0).sort_index(axis=1)
    return mat.reset_index().rename_axis(None, axis=1)


def _switchpoint_position_density(sw: pd.DataFrame, genome_summary: pd.DataFrame, bins: int = 120) -> pd.DataFrame:
    gs = genome_summary.copy()
    gs["name"] = gs["name"].astype(str).map(_tip_to_accession)
    lens = gs[["name", "length"]].rename(columns={"length": "genome_len"})
    tmp = sw.merge(lens, on="name", how="left")
    if tmp["genome_len"].isna().any():
        miss = sorted(tmp.loc[tmp["genome_len"].isna(), "name"].unique().tolist())
        raise ValueError(f"Missing genome lengths for switchpoints: {miss[:10]}")

    pos_norm = tmp["pos0"].astype(float) / tmp["genome_len"].astype(float)
    pos_norm = pos_norm.clip(lower=0.0, upper=np.nextafter(1.0, 0.0))
    hist, edges = np.histogram(pos_norm.to_numpy(), bins=bins, range=(0.0, 1.0))
    total = hist.sum()
    df = pd.DataFrame(
        {
            "bin_start": edges[:-1],
            "bin_end": edges[1:],
            "count": hist,
            "density": (hist / total) if total else np.nan,
        }
    )
    df["bin_mid"] = (df["bin_start"] + df["bin_end"]) / 2.0
    return df


def make_figure2(
    genome_summary_csv: Path,
    switchpoints_csv: Path,
    switchpoint_post_dir: Optional[Path],
    out_dir: Path,
    panel_case: str = "upper",
    dpi: int = 300,
    bins: int = 120,
) -> Tuple[Path, Path]:
    """
    Figure 2:
      (A) switchpoint transition matrix + standardized residual star overlay
      (B) normalized switchpoint position density
    """
    gs = pd.read_csv(genome_summary_csv)

    # Use postprocess outputs if present, else compute
    trans = None
    dens = None
    if switchpoint_post_dir:
        cand1 = switchpoint_post_dir / "switchpoint_transition_matrix.csv"
        cand2 = switchpoint_post_dir / "switchpoint_position_density.csv"
        if cand1.exists() and cand2.exists():
            trans = pd.read_csv(cand1)
            dens = pd.read_csv(cand2)

    if trans is None or dens is None:
        sw = pd.read_csv(switchpoints_csv)
        sw = _normalize_switchpoints(sw)
        trans = _switchpoint_transition_matrix(sw)
        dens = _switchpoint_position_density(sw, gs, bins=bins)

    mat = trans.set_index("from_label")
    mat_vals = mat.to_numpy(dtype=float)
    xlabels = list(mat.columns)
    ylabels = list(mat.index)

    fig = plt.figure(figsize=(13, 5.5))
    grid = fig.add_gridspec(nrows=1, ncols=2, width_ratios=[1.15, 1.0])
    axA = fig.add_subplot(grid[0, 0])
    axB = fig.add_subplot(grid[0, 1])

    im = axA.imshow(mat_vals, aspect="auto", interpolation="nearest")
    axA.set_title("Switchpoint transition matrix")
    axA.set_xlabel("To label")
    axA.set_ylabel("From label")
    axA.set_xticks(range(len(xlabels)))
    axA.set_xticklabels([str(x) for x in xlabels], rotation=90, fontsize=7)
    axA.set_yticks(range(len(ylabels)))
    axA.set_yticklabels([str(y) for y in ylabels], fontsize=7)
    cbar = fig.colorbar(im, ax=axA, fraction=0.046, pad=0.02)
    cbar.set_label("Count")

    # --- NEW: standardized residual overlay (* for |resid|>=3) ---
    M = mat_vals.astype(float)
    row_sum = M.sum(axis=1, keepdims=True)
    col_sum = M.sum(axis=0, keepdims=True)
    total = M.sum()
    if total > 0:
        expected = (row_sum @ col_sum) / total
        with np.errstate(divide="ignore", invalid="ignore"):
            resid = (M - expected) / np.sqrt(expected)
        for i in range(M.shape[0]):
            for j in range(M.shape[1]):
                if np.isfinite(resid[i, j]) and abs(resid[i, j]) >= 3.0:
                    axA.text(j, i, "*", ha="center", va="center", fontsize=10, fontweight="bold")
    # -------------------------------------------------------------

    _panel_label(axA, "A", case=panel_case)

    axB.plot(dens["bin_mid"], dens["density"])
    axB.set_title("Normalized switchpoint positional density")
    axB.set_xlabel("Genome position (normalized 0–1)")
    axB.set_ylabel("Density")
    _panel_label(axB, "B", case=panel_case)

    fig.tight_layout()
    out_pdf = out_dir / "Figure2_switchpoints.pdf"
    out_png = out_dir / "Figure2_switchpoints.png"
    _save_fig(fig, out_pdf, out_png, dpi=dpi)
    return out_pdf, out_png


def make_figure3(
    switchpoint_orf_distances_csv: Path,
    enrichment_summary_csv: Path,
    null_fracs_npy: Path,
    out_dir: Path,
    panel_case: str = "upper",
    dpi: int = 300,
    near_bp_default: int = 200,
) -> Tuple[Path, Path]:
    """
    Figure 3:
      (A) null distribution of fraction within ±near_bp of ORF boundary vs observed (textbox at top-right)
      (B) histogram of distances
      (C) CDF of distances
    """
    dist_df = pd.read_csv(switchpoint_orf_distances_csv)

    # pick distance column robustly
    dist_col = None
    for c in dist_df.columns:
        lc = c.lower()
        if "dist" in lc and ("nearest" in lc or "boundary" in lc):
            dist_col = c
            break
    if dist_col is None:
        for c in dist_df.columns:
            if "dist" in c.lower():
                dist_col = c
                break
    if dist_col is None:
        raise ValueError(f"Cannot find distance column in {switchpoint_orf_distances_csv}")

    d = pd.to_numeric(dist_df[dist_col], errors="coerce").dropna().to_numpy(dtype=float)
    d_abs = np.abs(d)

    null_fracs = np.load(null_fracs_npy)
    null_fracs = np.asarray(null_fracs, dtype=float)
    null_fracs = null_fracs[np.isfinite(null_fracs)]

    summ = pd.read_csv(enrichment_summary_csv)
    row = summ.iloc[0].to_dict() if len(summ) else {}

    near_bp = int(row.get("near_bp", near_bp_default))
    obs_frac = row.get("observed_frac", None)
    if obs_frac is None:
        obs_frac = float(np.mean(d_abs <= near_bp)) if d_abs.size else float("nan")
    else:
        obs_frac = float(obs_frac)

    pval = row.get("p_two_sided", row.get("p", None))
    if pval is None and null_fracs.size:
        pval = (1.0 + float(np.sum(null_fracs >= obs_frac))) / (len(null_fracs) + 1.0)
    if pval is not None:
        pval = float(pval)

    fig = plt.figure(figsize=(14, 4.8))
    grid = fig.add_gridspec(nrows=1, ncols=3, width_ratios=[1.0, 1.0, 1.0])
    axA = fig.add_subplot(grid[0, 0])
    axB = fig.add_subplot(grid[0, 1])
    axC = fig.add_subplot(grid[0, 2])

    # (A) null frac
    axA.hist(null_fracs, bins=30)
    axA.axvline(obs_frac, linestyle="--", linewidth=2)
    axA.set_title(r"Null: fraction within $\pm$ window of ORF boundary")
    axA.set_xlabel("Fraction")
    axA.set_ylabel("Count")
    # textbox at top-right (avoid overlap)
    txt = rf"Observed={obs_frac:.3g}" + "\n" + rf"$\pm {near_bp}$ bp"
    if pval is not None:
        txt += "\n" + rf"$p={pval:.3g}$"
    axA.text(
        0.98, 0.98, txt,
        transform=axA.transAxes,
        ha="right", va="top", fontsize=9,
        bbox=dict(boxstyle="round", alpha=0.2),
    )
    _panel_label(axA, "A", case=panel_case)

    # (B) histogram
    axB.hist(d_abs, bins=40)
    axB.axvline(near_bp, linestyle="--", linewidth=2)
    axB.set_title("Distance to nearest ORF boundary")
    axB.set_xlabel("Absolute distance (bp)")
    axB.set_ylabel("Count")
    _panel_label(axB, "B", case=panel_case)

    # (C) CDF
    if d_abs.size:
        xs = np.sort(d_abs)
        ys = np.arange(1, len(xs) + 1) / float(len(xs))
        axC.plot(xs, ys)
        axC.axvline(near_bp, linestyle="--", linewidth=2)
    axC.set_title("CDF: distance to nearest ORF boundary")
    axC.set_xlabel("Absolute distance (bp)")
    axC.set_ylabel("CDF")
    _panel_label(axC, "C", case=panel_case)

    fig.tight_layout()
    out_pdf = out_dir / "Figure3_switchpoint_orf_null.pdf"
    out_png = out_dir / "Figure3_switchpoint_orf_null.png"
    _save_fig(fig, out_pdf, out_png, dpi=dpi)
    return out_pdf, out_png


def make_figure4(
    orf_dist_csv: Path,
    out_dir: Path,
    panel_case: str = "upper",
    dpi: int = 300,
) -> Tuple[Path, Path]:
    """
    Figure 4 (deduplicated):
      ONLY ORF3 distance heatmap (tree is shown in Fig6B).
    """
    D = _read_square_matrix_csv(orf_dist_csv)
    D.index = [_tip_to_accession(x) for x in D.index]
    D.columns = [_tip_to_accession(x) for x in D.columns]

    fig = plt.figure(figsize=(10.5, 8.5))
    ax = fig.add_subplot(111)
    mat = D.to_numpy(dtype=float)
    im = ax.imshow(mat, aspect="equal", interpolation="nearest")
    ax.set_title("ORF3 pairwise identity distance")
    ax.set_xlabel("Genomes")
    ax.set_ylabel("Genomes")
    ax.set_xticks(range(len(D.columns)))
    ax.set_yticks(range(len(D.index)))
    ax.set_xticklabels(D.columns, rotation=90, fontsize=4)
    ax.set_yticklabels(D.index, fontsize=4)
    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.02)
    cbar.set_label("Distance")
    _panel_label(ax, "A", case=panel_case)

    fig.tight_layout()
    out_pdf = out_dir / "Figure4_orf3_distance_heatmap.pdf"
    out_png = out_dir / "Figure4_orf3_distance_heatmap.png"
    _save_fig(fig, out_pdf, out_png, dpi=dpi)
    return out_pdf, out_png


def make_figure5(
    kmer_dist_csv: Path,
    orf_dist_csv: Path,
    compare_summary_csv: Path,
    out_dir: Path,
    panel_case: str = "upper",
    dpi: int = 300,
    b_split: str = "|",
) -> Tuple[Path, Path]:
    """
    Figure 5:
      distance–distance scatter with italic mathtext for k, r, p
    """
    A = _read_square_matrix_csv(kmer_dist_csv)
    B = _read_square_matrix_csv(orf_dist_csv)

    A.index = [_tip_to_accession(x) for x in A.index]
    A.columns = [_tip_to_accession(x) for x in A.columns]
    if b_split:
        B.index = [str(x).split(b_split)[0] for x in B.index]
        B.columns = [str(x).split(b_split)[0] for x in B.columns]
    B.index = [_tip_to_accession(x) for x in B.index]
    B.columns = [_tip_to_accession(x) for x in B.columns]

    common = sorted(set(A.index) & set(B.index))
    if len(common) < 3:
        raise ValueError(f"Too few common genomes between matrices: {len(common)}")

    A = A.loc[common, common]
    B = B.loc[common, common]

    a = A.to_numpy(dtype=float)
    b = B.to_numpy(dtype=float)
    np.fill_diagonal(a, 0.0)
    np.fill_diagonal(b, 0.0)

    va = _upper_tri_vec(a)
    vb = _upper_tri_vec(b)

    summ = pd.read_csv(compare_summary_csv)
    row = summ.iloc[0].to_dict() if len(summ) else {}
    r = float(row.get("r", np.nan))
    p = float(row.get("p_two_sided", row.get("p", np.nan)))
    n_pairs = int(row.get("n_pairs", len(va)))

    fig = plt.figure(figsize=(7.0, 6.2))
    ax = fig.add_subplot(111)
    ax.scatter(va, vb, s=16, alpha=0.8)
    ax.set_xlabel(r"Genome $k$-mer cosine distance ($k=4$)")
    ax.set_ylabel("ORF3 distance (protein alignment)")
    ax.set_title(rf"Spearman $r$={r:.3f}, $p$={p:.4g}, $k=4$, n={n_pairs} pairs")
    _panel_label(ax, "A", case=panel_case)

    fig.tight_layout()
    out_pdf = out_dir / "Figure5_distance_concordance.pdf"
    out_png = out_dir / "Figure5_distance_concordance.png"
    _save_fig(fig, out_pdf, out_png, dpi=dpi)
    return out_pdf, out_png


def _parse_window_assignments(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    cols = {c.lower(): c for c in df.columns}

    if "genome" in cols:
        gcol = cols["genome"]
    elif "name" in cols:
        gcol = cols["name"]
    else:
        raise ValueError("window_assignments missing genome/name column")

    if "start" not in cols:
        raise ValueError("window_assignments missing start column")
    scol = cols["start"]

    if "label" in cols:
        lcol = cols["label"]
    elif "assigned_cluster" in cols:
        lcol = cols["assigned_cluster"]
    elif "cluster" in cols:
        lcol = cols["cluster"]
    else:
        raise ValueError("window_assignments missing label/assigned_cluster/cluster column")

    out = df[[gcol, scol, lcol]].copy()
    out.columns = ["genome", "start", "label"]
    out["genome"] = out["genome"].astype(str).map(_tip_to_accession)
    out["start"] = pd.to_numeric(out["start"], errors="coerce").fillna(0).astype(int)
    out["label"] = pd.to_numeric(out["label"], errors="coerce").fillna(0).astype(int)
    return out


def _draw_barcode(ax: plt.Axes, win: pd.DataFrame, genomes: List[str]) -> None:
    cmap = plt.get_cmap("tab20")
    g_to_y = {g: i for i, g in enumerate(genomes)}
    sub = win[win["genome"].isin(genomes)].copy().sort_values(["genome", "start"])
    xs = sub["start"].to_numpy()
    ys = sub["genome"].map(g_to_y).astype(int).to_numpy()
    cs = sub["label"].to_numpy()
    sc = ax.scatter(xs, ys, c=cs, cmap=cmap, marker="s", s=18)
    ax.set_yticks(range(len(genomes)))
    ax.set_yticklabels(genomes, fontsize=7)
    ax.set_xlabel("Genome position (start bp)")
    ax.set_title("Mosaic barcodes (top chimera candidates)")
    cb = plt.colorbar(sc, ax=ax, fraction=0.046, pad=0.02)
    cb.set_label("Mosaic label")


def _draw_tree_highlight(ax: plt.Axes, newick: Path, highlight: set[str]) -> None:
    from Bio import Phylo
    tree = Phylo.read(str(newick), "newick")
    for term in tree.get_terminals():
        acc = _tip_to_accession(term.name)
        term.name = f"{acc}*" if acc in highlight else acc
    Phylo.draw(tree, axes=ax, do_show=False)
    ax.set_title("ORF3 NJ tree (candidates in red, *)")
    ax.set_axis_off()
    for txt in ax.texts:
        t = txt.get_text().strip()
        acc = t[:-1] if t.endswith("*") else t
        txt.set_fontsize(7)
        if acc in highlight:
            txt.set_color("red")
            txt.set_fontweight("bold")


def make_figure6(
    chimera_candidates_csv: Path,
    window_assignments_csv: Path,
    tree_newick: Path,
    out_dir: Path,
    top_n: int = 10,
    panel_case: str = "upper",
    dpi: int = 300,
) -> Tuple[Path, Path]:
    """
    Figure 6:
      (A) barcodes for top N
      (B) ORF3 NJ tree with highlighted tips
    """
    chim = pd.read_csv(chimera_candidates_csv)
    if "genome" not in chim.columns:
        raise ValueError("chimera_candidates.csv must have 'genome' column")
    chim["genome"] = chim["genome"].astype(str).map(_tip_to_accession)
    top = chim.head(int(top_n))["genome"].tolist()
    highlight = set(top)

    win = _parse_window_assignments(window_assignments_csv)

    fig = plt.figure(figsize=(14, 0.35 * len(top) + 6))
    grid = fig.add_gridspec(nrows=1, ncols=2, width_ratios=[1.35, 1.0])

    axA = fig.add_subplot(grid[0, 0])
    _draw_barcode(axA, win, top)
    _panel_label(axA, "A", case=panel_case)

    axB = fig.add_subplot(grid[0, 1])
    _draw_tree_highlight(axB, tree_newick, highlight)
    _panel_label(axB, "B", case=panel_case)

    fig.tight_layout()
    out_pdf = out_dir / "Figure6_chimera_barcode_tree.pdf"
    out_png = out_dir / "Figure6_chimera_barcode_tree.png"
    _save_fig(fig, out_pdf, out_png, dpi=dpi)
    return out_pdf, out_png


# --------------------------
# Table 1 (Excel)
# --------------------------
def make_table1_excel(
    chimera_candidates_csv: Path,
    merged_csv: Path,
    out_dir: Path,
    top_n: int = 10,
) -> Tuple[Path, Path]:
    """
    Create Table 1 (top N chimera candidates) as:
      - Excel .xlsx
      - CSV
    """
    chim = pd.read_csv(chimera_candidates_csv)
    merged = pd.read_csv(merged_csv)
    if "genome" not in chim.columns or "genome" not in merged.columns:
        raise ValueError("chimera_candidates.csv and merged.csv must both have 'genome' column")

    chim["genome"] = chim["genome"].astype(str).map(_tip_to_accession)
    merged["genome"] = merged["genome"].astype(str).map(_tip_to_accession)

    top = chim.head(int(top_n))["genome"].tolist()
    rank_map = {g: i + 1 for i, g in enumerate(top)}

    df = merged[merged["genome"].isin(top)].copy()
    df["rank"] = df["genome"].map(rank_map).astype(int)
    df = df.sort_values("rank")

    want = [
        "rank",
        "genome",
        "chimera_score",
        "dominant_label",
        "dominant_frac",
        "n_labels",
        "label_entropy",
        "switch_rate",
        "n_switches",
        "orf_cluster",
        "mapped_mosaic_from_orf",
        "mismatch",
    ]
    cols = [c for c in want if c in df.columns]
    df_out = df[cols].copy()

    _ensure_dir(out_dir)
    xlsx_path = out_dir / f"Table1_chimera_top{top_n}.xlsx"
    csv_path = out_dir / f"Table1_chimera_top{top_n}.csv"
    df_out.to_csv(csv_path, index=False)

    import openpyxl
    from openpyxl.styles import Font, Alignment

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        df_out.to_excel(w, sheet_name="Table1", index=False)
        ws = w.book["Table1"]
        ws.freeze_panes = "A2"

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_widths = {c: max(10, min(22, len(c) + 2)) for c in df_out.columns}
        col_widths.update({"rank": 6, "genome": 14})

        for j, h in enumerate(df_out.columns.tolist(), start=1):
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = col_widths.get(h, 12)

        for i in range(2, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                ws.cell(row=i, column=j).alignment = Alignment(horizontal="center", vertical="center")

    return xlsx_path, csv_path


# --------------------------
# Supplementary Data S1 (single Excel with multi-sheets)
# --------------------------
def make_supplementary_s1_excel(
    gb_dir: Path,
    mosaic_orf_dir: Path,
    orf3_dir: Path,
    orf3_phylogeny_dir: Path,
    compare_dir: Path,
    agreement_dir: Path,
    switchpoint_post_dir: Optional[Path],
    figures_dir: Path,
    tables_dir: Path,
    out_xlsx: Path,
) -> Path:
    """
    Create Supplementary_Data_S1.xlsx with multiple sheets.

    Sheets include:
    - genome_summary, genome_embedding
    - k4_cosine_distance, window_assignments
    - predicted_orfs, switchpoints, switchpoint_orf_distances, enrichment_summary
    - orf3_pairwise_identity_distance
    - compare_distances_summary, agreement_metrics, contingency, chimera_candidates, mosaic_orf_merged
    - Table1 (final)
    - manifest (what was found/missing)
    """
    manifest_rows: List[Dict[str, str]] = []

    def add_sheet(writer, sheet_name: str, df: Optional[pd.DataFrame], src: Path):
        if df is None:
            manifest_rows.append({"sheet": sheet_name, "status": "MISSING", "source": str(src)})
            return
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        manifest_rows.append({"sheet": sheet_name, "status": "OK", "source": str(src)})

    def try_csv(path: Path) -> Optional[pd.DataFrame]:
        if not path.exists():
            return None
        try:
            if path.suffix.lower() == ".csv":
                if path.name.endswith(".csv") and path.stat().st_size > 0:
                    # distance matrices may be square index_col=0; keep both as dataframes
                    return pd.read_csv(path)
            return pd.read_csv(path)
        except Exception:
            return None

    # Load key tables
    genome_summary = try_csv(gb_dir / "genome_summary.csv")
    genome_embedding = try_csv(gb_dir / "genome_embedding.csv")
    window_assignments = try_csv(gb_dir / "window_assignments.csv")
    k4_cos = None
    if (gb_dir / "k4_cosine_distance.csv").exists():
        k4_cos = pd.read_csv(gb_dir / "k4_cosine_distance.csv", index_col=0).reset_index().rename(columns={"index": "name"})
    else:
        # fallback for naming variants
        for p in gb_dir.glob("k*_cosine_distance.csv"):
            k4_cos = pd.read_csv(p, index_col=0).reset_index().rename(columns={"index": "name"})
            break

    predicted_orfs = try_csv(mosaic_orf_dir / "predicted_orfs.csv")
    switchpoints = try_csv(mosaic_orf_dir / "switchpoints.csv")
    sp_orf_dist = try_csv(mosaic_orf_dir / "switchpoint_orf_distances.csv")
    enrichment_summary = try_csv(mosaic_orf_dir / "enrichment_summary.csv")

    orf3_dist = None
    if (orf3_phylogeny_dir / "pairwise_identity_distance.csv").exists():
        orf3_dist = pd.read_csv(orf3_phylogeny_dir / "pairwise_identity_distance.csv", index_col=0).reset_index().rename(columns={"index": "name"})

    compare_summary = try_csv(compare_dir / "compare_distances.summary.csv")
    agreement_metrics = try_csv(agreement_dir / "agreement_metrics.csv")
    contingency = try_csv(agreement_dir / "contingency_orf_vs_mosaic.csv")
    chimera_candidates = try_csv(agreement_dir / "chimera_candidates.csv")
    mosaic_orf_merged = try_csv(agreement_dir / "mosaic_orf_merged_per_genome.csv")

    # Optional postprocess
    sp_per_genome = try_csv(switchpoint_post_dir / "switchpoints_per_genome.csv") if switchpoint_post_dir else None
    sp_trans_mat = try_csv(switchpoint_post_dir / "switchpoint_transition_matrix.csv") if switchpoint_post_dir else None
    sp_pos_dens = try_csv(switchpoint_post_dir / "switchpoint_position_density.csv") if switchpoint_post_dir else None

    # Final Table1
    table1_xlsx = tables_dir / "Table1_chimera_top10.xlsx"
    table1_df = None
    if table1_xlsx.exists():
        try:
            table1_df = pd.read_excel(table1_xlsx, sheet_name="Table1")
        except Exception:
            table1_df = None

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        add_sheet(writer, "genome_summary", genome_summary, gb_dir / "genome_summary.csv")
        add_sheet(writer, "genome_embedding", genome_embedding, gb_dir / "genome_embedding.csv")
        add_sheet(writer, "k4_cosine_distance", k4_cos, gb_dir / "k4_cosine_distance.csv")
        add_sheet(writer, "window_assignments", window_assignments, gb_dir / "window_assignments.csv")

        add_sheet(writer, "predicted_orfs", predicted_orfs, mosaic_orf_dir / "predicted_orfs.csv")
        add_sheet(writer, "switchpoints", switchpoints, mosaic_orf_dir / "switchpoints.csv")
        add_sheet(writer, "switchpoint_orf_distances", sp_orf_dist, mosaic_orf_dir / "switchpoint_orf_distances.csv")
        add_sheet(writer, "enrichment_summary", enrichment_summary, mosaic_orf_dir / "enrichment_summary.csv")

        add_sheet(writer, "orf3_distance", orf3_dist, orf3_phylogeny_dir / "pairwise_identity_distance.csv")
        add_sheet(writer, "compare_distances_summary", compare_summary, compare_dir / "compare_distances.summary.csv")
        add_sheet(writer, "agreement_metrics", agreement_metrics, agreement_dir / "agreement_metrics.csv")
        add_sheet(writer, "contingency_orf_vs_mosaic", contingency, agreement_dir / "contingency_orf_vs_mosaic.csv")
        add_sheet(writer, "chimera_candidates", chimera_candidates, agreement_dir / "chimera_candidates.csv")
        add_sheet(writer, "mosaic_orf_merged", mosaic_orf_merged, agreement_dir / "mosaic_orf_merged_per_genome.csv")

        add_sheet(writer, "switchpoints_per_genome", sp_per_genome, (switchpoint_post_dir / "switchpoints_per_genome.csv") if switchpoint_post_dir else Path("N/A"))
        add_sheet(writer, "switchpoint_transition_matrix", sp_trans_mat, (switchpoint_post_dir / "switchpoint_transition_matrix.csv") if switchpoint_post_dir else Path("N/A"))
        add_sheet(writer, "switchpoint_position_density", sp_pos_dens, (switchpoint_post_dir / "switchpoint_position_density.csv") if switchpoint_post_dir else Path("N/A"))

        add_sheet(writer, "Table1", table1_df, table1_xlsx)

        manifest_df = pd.DataFrame(manifest_rows)
        manifest_df.to_excel(writer, sheet_name="manifest", index=False)

    return out_xlsx


# --------------------------
# Main
# --------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--gb_dir", required=True)
    ap.add_argument("--mosaic_orf_dir", required=True)
    ap.add_argument("--orf3_dir", required=True)
    ap.add_argument("--orf3_phylogeny_dir", required=True)
    ap.add_argument("--compare_dir", required=True)
    ap.add_argument("--agreement_dir", required=True)
    ap.add_argument("--switchpoint_post_dir", default=None)

    ap.add_argument("--out_dir", required=True)
    ap.add_argument("--top_n", type=int, default=10)
    ap.add_argument("--panel_case", choices=["upper", "lower"], default="upper")
    ap.add_argument("--dpi", type=int, default=300)

    ap.add_argument("--skip_figures", action="store_true")
    ap.add_argument("--skip_table", action="store_true")
    ap.add_argument("--skip_s1", action="store_true")
    args = ap.parse_args()

    gb_dir = Path(args.gb_dir)
    mosaic_orf_dir = Path(args.mosaic_orf_dir)
    orf3_dir = Path(args.orf3_dir)
    orf3_phylogeny_dir = Path(args.orf3_phylogeny_dir)
    compare_dir = Path(args.compare_dir)
    agreement_dir = Path(args.agreement_dir)
    switchpoint_post_dir = Path(args.switchpoint_post_dir) if args.switchpoint_post_dir else None

    out_root = Path(args.out_dir)
    figures_dir = _ensure_dir(out_root / "figures")
    tables_dir = _ensure_dir(out_root / "tables")
    supp_dir = _ensure_dir(out_root / "supplementary")

    # Figures
    if not args.skip_figures:
        make_figure1(
            genome_summary_csv=gb_dir / "genome_summary.csv",
            genome_embedding_csv=gb_dir / "genome_embedding.csv",
            kmer_dist_csv=gb_dir / "k4_cosine_distance.csv",
            out_dir=figures_dir,
            panel_case=args.panel_case,
            dpi=args.dpi,
        )

        make_figure2(
            genome_summary_csv=gb_dir / "genome_summary.csv",
            switchpoints_csv=mosaic_orf_dir / "switchpoints.csv",
            switchpoint_post_dir=switchpoint_post_dir,
            out_dir=figures_dir,
            panel_case=args.panel_case,
            dpi=args.dpi,
            bins=120,
        )

        make_figure3(
            switchpoint_orf_distances_csv=mosaic_orf_dir / "switchpoint_orf_distances.csv",
            enrichment_summary_csv=mosaic_orf_dir / "enrichment_summary.csv",
            null_fracs_npy=mosaic_orf_dir / "null_fracs.npy",
            out_dir=figures_dir,
            panel_case=args.panel_case,
            dpi=args.dpi,
            near_bp_default=200,
        )

        make_figure4(
            orf_dist_csv=orf3_phylogeny_dir / "pairwise_identity_distance.csv",
            out_dir=figures_dir,
            panel_case=args.panel_case,
            dpi=args.dpi,
        )

        make_figure5(
            kmer_dist_csv=gb_dir / "k4_cosine_distance.csv",
            orf_dist_csv=orf3_phylogeny_dir / "pairwise_identity_distance.csv",
            compare_summary_csv=compare_dir / "compare_distances.summary.csv",
            out_dir=figures_dir,
            panel_case=args.panel_case,
            dpi=args.dpi,
            b_split="|",
        )

        make_figure6(
            chimera_candidates_csv=agreement_dir / "chimera_candidates.csv",
            window_assignments_csv=gb_dir / "window_assignments.csv",
            tree_newick=orf3_phylogeny_dir / "nj_tree.newick",
            out_dir=figures_dir,
            top_n=args.top_n,
            panel_case=args.panel_case,
            dpi=args.dpi,
        )

    # Table 1
    if not args.skip_table:
        make_table1_excel(
            chimera_candidates_csv=agreement_dir / "chimera_candidates.csv",
            merged_csv=agreement_dir / "mosaic_orf_merged_per_genome.csv",
            out_dir=tables_dir,
            top_n=args.top_n,
        )

    # Supplementary Data S1 as multi-sheet Excel
    if not args.skip_s1:
        out_xlsx = supp_dir / "Supplementary_Data_S1.xlsx"
        make_supplementary_s1_excel(
            gb_dir=gb_dir,
            mosaic_orf_dir=mosaic_orf_dir,
            orf3_dir=orf3_dir,
            orf3_phylogeny_dir=orf3_phylogeny_dir,
            compare_dir=compare_dir,
            agreement_dir=agreement_dir,
            switchpoint_post_dir=switchpoint_post_dir,
            figures_dir=figures_dir,
            tables_dir=tables_dir,
            out_xlsx=out_xlsx,
        )
        print(f"[OK] Supplementary Data S1 Excel: {out_xlsx}")

    print("Done.")
    print(f"- Figures: {figures_dir}")
    print(f"- Tables:  {tables_dir}")
    print(f"- Supp:    {supp_dir}")


if __name__ == "__main__":
    main()
